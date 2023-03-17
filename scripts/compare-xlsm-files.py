import openpyxl

def compare_files(file1_path, file2_path):


    #3 worksheets
    #1 before
    #2 after
    #3 verification 


    # Load both XLSM files
    file1 = openpyxl.load_workbook(filename=file1_path)
    file2 = openpyxl.load_workbook(filename=file2_path)

    # Get the first worksheet from both files
    sheet1 = file1.active
    sheet2 = file2.active

    # Loop through all rows in the first worksheet
    for row in range(1, sheet1.max_row + 1):
        # Loop through all columns in the row
        for col in range(1, sheet1.max_column + 1):
            # Get the value from the corresponding cell in the second worksheet
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            # Compare the values in the two cells
            if cell1.value != cell2.value:
                # If the values don't match, return False
                return False

    # If all values match, return True
    return True

    